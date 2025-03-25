import pandas as pd
import os
import logging
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.conf import settings

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Service to generate Excel reports from call analysis data."""
    
    def generate_call_report(self, call_analysis):
        """
        Generate a detailed Excel report for a single call analysis.
        
        Args:
            call_analysis: CallAnalysis object containing the analysis data
            
        Returns:
            str: Path to the generated Excel file
        """
        try:
            logger.info(f"Generating report for call: {call_analysis.call_recording.title}")
            
            # Create a dataframe with basic call info
            basic_info = {
                'Call Title': [call_analysis.call_recording.title],
                'Agent': [call_analysis.agent.user.get_full_name()],
                'Date': [call_analysis.call_recording.uploaded_at.strftime('%Y-%m-%d')],
                'Time': [call_analysis.call_recording.uploaded_at.strftime('%H:%M:%S')],
                'Duration (seconds)': [call_analysis.call_recording.duration_seconds],
                'Coverage Score': [call_analysis.coverage_score],
                'Sentiment': [call_analysis.sentiment]
            }
            
            basic_df = pd.DataFrame(basic_info)
            
            # Create a dataframe for the transcription
            utterances = []
            for utterance in call_analysis.transcription_text.split('\n'):
                if utterance.strip():
                    utterances.append(utterance)
            
            transcript_df = pd.DataFrame({'Transcription': utterances})
            
            # Create dataframe for analysis results
            key_issues = call_analysis.key_issues
            if isinstance(key_issues, list):
                key_issues_str = ', '.join(key_issues)
            else:
                key_issues_str = str(key_issues)
                
            analysis_data = {
                'Category': [
                    'Sentiment', 
                    'Coverage Score', 
                    'Score Explanation', 
                    'Key Issues', 
                    'Improvement Suggestions'
                ],
                'Value': [
                    call_analysis.sentiment,
                    str(call_analysis.coverage_score),
                    call_analysis.score_explanation,
                    key_issues_str,
                    call_analysis.improvement_suggestions
                ]
            }
            
            analysis_df = pd.DataFrame(analysis_data)
            
            # Generate file path
            reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
            os.makedirs(reports_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_name = f"call_report_{call_analysis.id}_{timestamp}.xlsx"
            file_path = os.path.join(reports_dir, file_name)
            
            # Create Excel writer
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Write basic info
                basic_df.to_excel(writer, sheet_name='Call Summary', index=False, startrow=1)
                
                # Write transcript
                transcript_df.to_excel(writer, sheet_name='Transcription', index=False, startrow=1)
                
                # Write analysis
                analysis_df.to_excel(writer, sheet_name='Analysis', index=False, startrow=1)
                
                # Get workbook and apply formatting
                workbook = writer.book
                
                # Format Call Summary sheet
                summary_sheet = workbook['Call Summary']
                self._apply_header_formatting(summary_sheet)
                
                # Format Transcription sheet
                transcript_sheet = workbook['Transcription']
                self._apply_header_formatting(transcript_sheet)
                
                # Format Analysis sheet
                analysis_sheet = workbook['Analysis']
                self._apply_header_formatting(analysis_sheet)
                
                # Auto-adjust column widths
                for sheet in workbook.worksheets:
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        adjusted_width = (max_length + 2)
                        sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
            
            logger.info(f"Report generated: {file_path}")
            return file_path
            
        except Exception as e:
            logger.exception(f"Exception in report generation: {str(e)}")
            return None
    
    def generate_aggregate_report(self, report_type, date_range_start, date_range_end, call_analyses):
        """
        Generate an aggregate report for multiple call analyses.
        
        Args:
            report_type: Type of report (weekly, monthly, custom)
            date_range_start: Start date for the report
            date_range_end: End date for the report
            call_analyses: List of CallAnalysis objects
            
        Returns:
            str: Path to the generated Excel file
        """
        try:
            logger.info(f"Generating {report_type} report from {date_range_start} to {date_range_end}")
            
            # Create basic report data
            report_data = []
            for analysis in call_analyses:
                call_data = {
                    'Call ID': analysis.id,
                    'Call Title': analysis.call_recording.title,
                    'Agent': analysis.agent.user.get_full_name(),
                    'Date': analysis.call_recording.uploaded_at.strftime('%Y-%m-%d'),
                    'Duration (s)': analysis.call_recording.duration_seconds,
                    'Coverage Score': analysis.coverage_score,
                    'Sentiment': analysis.sentiment,
                    'Key Issues': ', '.join(analysis.key_issues) if isinstance(analysis.key_issues, list) else str(analysis.key_issues)
                }
                report_data.append(call_data)
            
            # Create DataFrame
            calls_df = pd.DataFrame(report_data)
            
            # Calculate agent performance metrics
            if not calls_df.empty:
                agent_metrics = calls_df.groupby('Agent').agg({
                    'Call ID': 'count',
                    'Coverage Score': 'mean',
                    'Duration (s)': 'mean'
                }).reset_index()
                
                agent_metrics.columns = ['Agent', 'Total Calls', 'Avg Score', 'Avg Duration (s)']
                agent_metrics = agent_metrics.sort_values('Avg Score', ascending=False)
                
                # Extract common issues
                all_issues = []
                for issues in calls_df['Key Issues']:
                    for issue in issues.split(', '):
                        if issue.strip():
                            all_issues.append(issue.strip())
                
                issues_count = pd.Series(all_issues).value_counts().reset_index()
                issues_count.columns = ['Issue', 'Count']
                issues_count = issues_count.sort_values('Count', ascending=False)
            else:
                # Create empty dataframes if no data
                agent_metrics = pd.DataFrame(columns=['Agent', 'Total Calls', 'Avg Score', 'Avg Duration (s)'])
                issues_count = pd.DataFrame(columns=['Issue', 'Count'])
            
            # Generate file path
            reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
            os.makedirs(reports_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_name = f"{report_type}_report_{timestamp}.xlsx"
            file_path = os.path.join(reports_dir, file_name)
            
            # Create Excel writer
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Write data
                calls_df.to_excel(writer, sheet_name='Calls', index=False, startrow=1)
                agent_metrics.to_excel(writer, sheet_name='Agent Performance', index=False, startrow=1)
                issues_count.to_excel(writer, sheet_name='Common Issues', index=False, startrow=1)
                
                # Get workbook and apply formatting
                workbook = writer.book
                
                # Format sheets
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    self._apply_header_formatting(sheet)
                    
                    # Auto-adjust column widths
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        adjusted_width = (max_length + 2)
                        sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
            
            logger.info(f"Aggregate report generated: {file_path}")
            return file_path
            
        except Exception as e:
            logger.exception(f"Exception in aggregate report generation: {str(e)}")
            return None
    
    def _apply_header_formatting(self, sheet):
        """Apply formatting to the header row of a sheet."""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Format header cells
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add border
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
